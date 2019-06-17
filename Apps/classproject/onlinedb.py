'''import click
from onlineapp.models import College
import openpyxl
import django
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
django.setup()


@click.group()
@click.pass_context
def sqlop(ctx):
    pass

@sqlop.command('importdb')
@click.pass_context
def importdb(ctx):
    wb1 = openpyxl.load_workbook("C:\work\summer2019\precourse\students.xlsx")
    #wb2 = openpyxl.load_workbook("marks.xlsx")
    sheet1 = wb1.worksheets[2]
    c = College()
    for i in range(2, sheet1.max_row+1):
        college_data = []
        for j in range(1, sheet1.max_column + 1):
            college_data.append(sheet1.cell(row=i, column=j).value)
        c = College(college_data[0],college_data[1],college_data[2],college_data[3])
        c.save()

    c.save()


if __name__ == "__main__":
    sqlop(obj={})'''

import click
from onlineapp.models import *
import openpyxl
import django
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
django.setup()


def importdb():
    wb1 = openpyxl.load_workbook("C:\work\summer2019\precourse\students.xlsx")
    #wb2 = openpyxl.load_workbook("marks.xlsx")
    sheet1 = wb1.worksheets[2]
    c = College()
    for i in range(2, sheet1.max_row+1):
        college_data = []
        for j in range(1, sheet1.max_column + 1):
            college_data.append(sheet1.cell(row=i, column=j).value)
        c = College(college_data[0],college_data[1],college_data[2],college_data[3])
        c.save()

    c.save()
importdb()