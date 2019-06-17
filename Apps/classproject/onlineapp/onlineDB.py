import click

import openpyxl

import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
import django
django.setup()
from onlineapp.models import *


def importCollege():
    wb1 = openpyxl.load_workbook("C:\work\summer2019\precourse\students.xlsx")
    sheet1 = wb1.worksheets[2]
    c = College()
    for i in range(2, sheet1.max_row+1):
        college_data = []
        for j in range(1, sheet1.max_column + 1):
            college_data.append(sheet1.cell(row=i, column=j).value)
        c = College(name=college_data[0],acronym=college_data[1],location=college_data[2],contact=college_data[3])
        c.save()
    c.save()

def importStudents():
    wb = openpyxl.load_workbook("C:\work\summer2019\precourse\students.xlsx")
    sheet = wb.worksheets[0]
    for i in range(2,sheet.max_row+1):
        student_data = []
        for j in range(1, sheet.max_column + 1):
            student_data.append(sheet.cell(row=i, column=j).value)
        student = Student(name=student_data[0], email=student_data[2], db_folder=student_data[3],
                          college=College.objects.get(acronym=student_data[1]))
        student.save()

    sheet1 = wb.worksheets[1]
    for i in range(2,sheet1.max_row+1):
        student_data = []
        for j in range(1, sheet1.max_column + 1):
            student_data.append(sheet.cell(row=i, column=j).value)
        student = Student(name=student_data[0], email=student_data[2], dropped_out = True,db_folder=student_data[3],
                          college=College.objects.get(acronym=student_data[1]))
        student.save()

def mockTest():
    wb = openpyxl.load_workbook("C:\work\summer2019\precourse\marks.xlsx")
    sheet = wb.worksheets[0]
    for i in range(2, sheet.max_row + 1):
        mock_data = []
        for j in range(1,sheet.max_column+1):
            mock_data.append(sheet.cell(row=i, column=j).value)
        clg_end_idx = mock_data[0][7:].index("_")
        college = mock_data[0][7:7 + clg_end_idx]
        name = mock_data[0][7 + clg_end_idx + 1:-5:]
        mock = MockTest1(problem1=mock_data[1], problem2=mock_data[2], problem3=mock_data[3],
                         problem4=mock_data[4], total=mock_data[5],
                         student=Student.objects.get(db_folder=name,dropped_out=False))
        mock.save()
    mock.save()
#importStudents()
mockTest()